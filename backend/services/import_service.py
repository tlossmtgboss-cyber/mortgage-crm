"""
Import Service - CSV/Excel imports with transaction safety, streaming, and rollback.

Provides:
  - Stream-based CSV/Excel parsing (never loads entire file into memory)
  - Savepoint-wrapped imports for atomic rollback on failure
  - Batch processing with configurable chunk size
  - Import audit trail via import_jobs table
  - Hard-delete and soft-delete rollback options
  - Thread-safe progress tracking for long-running imports

Used by: backend/routes/data_import_routes.py
"""

import csv
import io
import json
import logging
import threading
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, BinaryIO, Callable, Dict, Generator, List, Optional, Tuple

from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
BATCH_SIZE = 100  # Rows per savepoint batch
MAX_ROWS = 100_000  # Hard cap on rows per import
PREVIEW_ROW_COUNT = 10


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------

@dataclass
class ImportProgress:
    """Thread-safe progress tracker for an import job."""
    total_rows: int = 0
    processed_rows: int = 0
    imported_rows: int = 0
    skipped_rows: int = 0
    error_rows: int = 0
    updated_rows: int = 0
    status: str = "pending"
    errors: List[Dict[str, Any]] = field(default_factory=list)
    imported_ids: List[int] = field(default_factory=list)
    _lock: threading.Lock = field(default_factory=threading.Lock, repr=False)

    @property
    def percentage(self) -> float:
        if self.total_rows <= 0:
            return 0.0
        return round((self.processed_rows / self.total_rows) * 100, 1)

    def add_imported(self, record_id: int) -> None:
        with self._lock:
            self.imported_rows += 1
            self.processed_rows += 1
            self.imported_ids.append(record_id)

    def add_updated(self, record_id: int) -> None:
        with self._lock:
            self.updated_rows += 1
            self.processed_rows += 1
            self.imported_ids.append(record_id)

    def add_skipped(self) -> None:
        with self._lock:
            self.skipped_rows += 1
            self.processed_rows += 1

    def add_error(self, row_num: int, errors: List[str]) -> None:
        with self._lock:
            self.error_rows += 1
            self.processed_rows += 1
            if len(self.errors) < 200:  # Cap stored errors
                self.errors.append({"row": row_num, "errors": errors})


@dataclass
class ImportResult:
    """Final result of an import operation."""
    import_id: str
    status: str  # completed | failed | rolled_back
    filename: str
    total_rows: int
    imported: int
    updated: int
    skipped: int
    errors: int
    error_details: List[Dict[str, Any]]
    imported_ids: List[int]
    field_mapping: Dict[str, str]
    message: str = ""


# ---------------------------------------------------------------------------
# Streaming parsers
# ---------------------------------------------------------------------------

def stream_csv_rows(
    file_stream: BinaryIO,
    max_file_size: int = MAX_FILE_SIZE,
    max_rows: int = MAX_ROWS,
) -> Generator[Tuple[List[str], Dict[str, Any], int], None, None]:
    """Stream CSV rows one at a time without loading entire file into memory.

    Yields (headers, row_dict, row_number) tuples. The headers list is
    yielded with every row for convenience but is computed once.

    The file_stream is read in chunks via the TextIOWrapper; at no point
    is the full file buffered in Python.

    Raises ValueError if file exceeds max_file_size.
    """
    # Read a small sample to detect encoding and dialect
    sample = file_stream.read(8192)
    if not sample:
        raise ValueError("CSV file is empty")

    # Detect encoding from BOM or fall back to utf-8
    if sample[:3] == b"\xef\xbb\xbf":
        encoding = "utf-8-sig"
    else:
        encoding = "utf-8"

    # Try to decode sample; if that fails, try latin-1
    try:
        sample.decode(encoding)
    except UnicodeDecodeError:
        encoding = "latin-1"

    file_stream.seek(0)

    # Wrap binary stream in a text wrapper for line-by-line reading
    text_stream = io.TextIOWrapper(file_stream, encoding=encoding, errors="replace")

    # Sniff dialect from first chunk
    first_chunk = text_stream.read(8192)
    try:
        dialect = csv.Sniffer().sniff(first_chunk, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    text_stream.seek(0)

    # Track bytes read for size limit enforcement
    bytes_read = 0

    reader = csv.DictReader(text_stream, dialect=dialect)
    headers = reader.fieldnames or []

    if not headers:
        raise ValueError("No column headers detected in CSV file")

    row_count = 0
    for row in reader:
        row_count += 1
        if row_count > max_rows:
            logger.warning(f"CSV import capped at {max_rows} rows")
            break

        # Approximate size tracking (exact would require re-encoding)
        bytes_read += sum(len(str(v or "")) for v in row.values()) + len(row) * 2
        if bytes_read > max_file_size:
            raise ValueError(
                f"File exceeds maximum size of {max_file_size // (1024 * 1024)}MB "
                f"(stopped at row {row_count})"
            )

        yield headers, row, row_count


def stream_excel_rows(
    file_stream: BinaryIO,
    sheet_name: Optional[str] = None,
    max_rows: int = MAX_ROWS,
) -> Generator[Tuple[List[str], Dict[str, Any], int], None, None]:
    """Stream Excel rows using openpyxl read_only mode.

    Yields (headers, row_dict, row_number) tuples.

    read_only=True means openpyxl does not load the entire workbook
    into memory; it streams row data from the XML on demand.
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError(
            "Excel import requires the openpyxl package. "
            "Install it with: pip install openpyxl"
        )

    # openpyxl needs a seekable bytes buffer
    # If the stream is already a BytesIO, great. If it's a SpooledTemporaryFile
    # from FastAPI, it should also work directly. We do NOT read the whole file
    # into a new BytesIO — openpyxl's read_only mode handles streaming internally.
    wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    if ws is None:
        wb.close()
        raise ValueError("Excel file has no active worksheet")

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        wb.close()
        raise ValueError("Excel file is empty")

    headers = [
        str(h).strip() if h is not None else f"column_{i}"
        for i, h in enumerate(raw_headers)
    ]

    row_count = 0
    for row_values in rows_iter:
        # Skip completely empty rows
        if not any(v is not None and str(v).strip() for v in row_values):
            continue

        row_count += 1
        if row_count > max_rows:
            logger.warning(f"Excel import capped at {max_rows} rows")
            break

        row_dict = {}
        for j, val in enumerate(row_values):
            if j < len(headers):
                row_dict[headers[j]] = val

        yield headers, row_dict, row_count

    wb.close()


# ---------------------------------------------------------------------------
# File size checker
# ---------------------------------------------------------------------------

def check_file_size(file_stream: BinaryIO, max_size: int = MAX_FILE_SIZE) -> int:
    """Check file size without reading entire contents into memory.

    Seeks to end, reads position, seeks back to start.
    Returns size in bytes. Raises ValueError if too large.
    """
    file_stream.seek(0, 2)  # Seek to end
    size = file_stream.tell()
    file_stream.seek(0)  # Reset to start

    if size > max_size:
        raise ValueError(
            f"File size ({size / (1024 * 1024):.1f}MB) exceeds maximum "
            f"allowed size ({max_size / (1024 * 1024):.0f}MB)"
        )

    return size


# ---------------------------------------------------------------------------
# Import Service
# ---------------------------------------------------------------------------

class ImportService:
    """Handles CSV/Excel imports with transaction safety, streaming, and rollback.

    Key design decisions:
    - Entire import runs inside a single DB transaction
    - Each batch of BATCH_SIZE rows is wrapped in a savepoint
    - If a batch fails, only that batch is rolled back (not the whole import)
    - Progress is tracked in-memory via ImportProgress (not via periodic commits)
    - The import_jobs row is written AFTER the import completes (single commit)
    - Rollback support: hard-delete (DELETE) or soft-delete (stage='Withdrawn')
    """

    def __init__(
        self,
        db: Session,
        user_id: int,
        organization_id: int,
    ):
        self.db = db
        self.user_id = user_id
        self.organization_id = organization_id

        # In-flight progress objects keyed by import_id
        self._active_imports: Dict[str, ImportProgress] = {}

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def import_csv(
        self,
        file_stream: BinaryIO,
        entity_type: str,
        field_mapping: Dict[str, str],
        filename: str = "upload.csv",
        duplicate_strategy: str = "skip",
        default_stage: Optional[str] = "New",
        default_source: Optional[str] = None,
        skip_invalid_rows: bool = True,
        validate_row_fn: Optional[Callable] = None,
        transform_value_fn: Optional[Callable] = None,
    ) -> ImportResult:
        """Stream-process a CSV file with transaction safety.

        Args:
            file_stream: Binary file-like object (e.g., UploadFile.file)
            entity_type: Target table ('leads', 'loans', etc.)
            field_mapping: {source_column: target_field} mapping
            filename: Original filename for audit trail
            duplicate_strategy: 'skip' | 'update' | 'create'
            default_stage: Default stage for new records
            default_source: Default source value
            skip_invalid_rows: If False, abort on first validation error
            validate_row_fn: Optional (row_dict, row_num) -> (clean, errors)
            transform_value_fn: Optional (field, value) -> transformed_value
        """
        check_file_size(file_stream)

        row_generator = stream_csv_rows(file_stream)
        return self._execute_import(
            row_generator=row_generator,
            entity_type=entity_type,
            field_mapping=field_mapping,
            filename=filename,
            duplicate_strategy=duplicate_strategy,
            default_stage=default_stage,
            default_source=default_source,
            skip_invalid_rows=skip_invalid_rows,
            validate_row_fn=validate_row_fn,
            transform_value_fn=transform_value_fn,
        )

    def import_excel(
        self,
        file_stream: BinaryIO,
        entity_type: str,
        field_mapping: Dict[str, str],
        filename: str = "upload.xlsx",
        sheet_name: Optional[str] = None,
        duplicate_strategy: str = "skip",
        default_stage: Optional[str] = "New",
        default_source: Optional[str] = None,
        skip_invalid_rows: bool = True,
        validate_row_fn: Optional[Callable] = None,
        transform_value_fn: Optional[Callable] = None,
    ) -> ImportResult:
        """Stream-process an Excel file with transaction safety.

        Args:
            file_stream: Binary file-like object
            entity_type: Target table ('leads', 'loans', etc.)
            field_mapping: {source_column: target_field} mapping
            filename: Original filename for audit trail
            sheet_name: Specific sheet to import (None = active sheet)
            duplicate_strategy: 'skip' | 'update' | 'create'
            default_stage: Default stage for new records
            default_source: Default source value
            skip_invalid_rows: If False, abort on first validation error
            validate_row_fn: Optional (row_dict, row_num) -> (clean, errors)
            transform_value_fn: Optional (field, value) -> transformed_value
        """
        check_file_size(file_stream)

        row_generator = stream_excel_rows(file_stream, sheet_name=sheet_name)
        return self._execute_import(
            row_generator=row_generator,
            entity_type=entity_type,
            field_mapping=field_mapping,
            filename=filename,
            duplicate_strategy=duplicate_strategy,
            default_stage=default_stage,
            default_source=default_source,
            skip_invalid_rows=skip_invalid_rows,
            validate_row_fn=validate_row_fn,
            transform_value_fn=transform_value_fn,
        )

    def rollback_import(
        self,
        batch_id: str,
        hard_delete: bool = False,
    ) -> Dict[str, Any]:
        """Roll back all records created by a specific import batch.

        Args:
            batch_id: The import job ID to roll back
            hard_delete: If True, DELETE records. If False, set stage='Withdrawn'.

        Returns:
            Dict with rollback results
        """
        # Fetch the import job
        job = self.db.execute(text("""
            SELECT id, status, imported_lead_ids, rolled_back_at, filename
            FROM import_jobs
            WHERE id = :id AND organization_id = :org_id
        """), {"id": batch_id, "org_id": self.organization_id}).fetchone()

        if not job:
            raise ValueError(f"Import job {batch_id} not found")

        job_id, status, lead_ids_raw, rolled_back_at, filename = job

        if status != "completed":
            raise ValueError(
                f"Cannot rollback import with status '{status}'. "
                "Only 'completed' imports can be rolled back."
            )

        if rolled_back_at is not None:
            raise ValueError("This import has already been rolled back")

        # Parse lead IDs
        if isinstance(lead_ids_raw, str):
            try:
                lead_ids = json.loads(lead_ids_raw)
            except json.JSONDecodeError:
                lead_ids = []
        elif isinstance(lead_ids_raw, list):
            lead_ids = lead_ids_raw
        else:
            lead_ids = []

        if not lead_ids:
            raise ValueError(
                "No imported record IDs recorded for this import. "
                "Rollback not possible."
            )

        now = datetime.now(timezone.utc)
        rolled_back_count = 0

        # Use a savepoint so rollback of the rollback is possible
        savepoint = self.db.begin_nested()
        try:
            # Process in batches of 100 to avoid oversized IN clauses
            for batch_start in range(0, len(lead_ids), BATCH_SIZE):
                batch = lead_ids[batch_start:batch_start + BATCH_SIZE]
                placeholders = ", ".join(f":id_{j}" for j in range(len(batch)))
                params: Dict[str, Any] = {
                    f"id_{j}": lid for j, lid in enumerate(batch)
                }
                params["org_id"] = self.organization_id

                if hard_delete:
                    query = f"""
                        DELETE FROM leads
                        WHERE id IN ({placeholders})
                          AND organization_id = :org_id
                    """
                    result = self.db.execute(text(query), params)
                else:
                    rollback_note = (
                        f"[IMPORT ROLLBACK] Rolled back import {batch_id} "
                        f"at {now.isoformat()}"
                    )
                    params["note"] = rollback_note
                    query = f"""
                        UPDATE leads
                        SET stage = 'Withdrawn',
                            notes = CASE
                                WHEN notes IS NULL THEN :note
                                ELSE notes || E'\\n' || :note
                            END
                        WHERE id IN ({placeholders})
                          AND organization_id = :org_id
                    """
                    result = self.db.execute(text(query), params)

                rolled_back_count += result.rowcount

            # Update import job status
            self.db.execute(text("""
                UPDATE import_jobs
                SET status = 'rolled_back',
                    rolled_back_at = :now
                WHERE id = :id
            """), {"id": batch_id, "now": now})

            savepoint.commit()

        except Exception:
            savepoint.rollback()
            raise

        self.db.commit()

        return {
            "import_id": batch_id,
            "status": "rolled_back",
            "method": "hard_delete" if hard_delete else "soft_delete",
            "leads_affected": rolled_back_count,
            "total_lead_ids": len(lead_ids),
            "rolled_back_at": now.isoformat(),
            "message": (
                f"Successfully {'deleted' if hard_delete else 'withdrew'} "
                f"{rolled_back_count} leads"
            ),
        }

    def get_import_history(
        self,
        limit: int = 20,
        offset: int = 0,
        status_filter: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Get recent import history for the organization.

        Returns:
            Dict with total count and list of import summaries
        """
        filters = ["organization_id = :org_id"]
        params: Dict[str, Any] = {
            "org_id": self.organization_id,
            "limit": limit,
            "offset": offset,
        }

        if status_filter:
            filters.append("status = :status")
            params["status"] = status_filter

        where_clause = " AND ".join(filters)

        query = f"""
            SELECT COUNT(*) FROM import_jobs WHERE {where_clause}
        """
        count_row = self.db.execute(text(query), params).fetchone()
        total = count_row[0] if count_row else 0

        query = f"""
            SELECT id, status, filename, total_rows, imported_rows,
                   skipped_rows, error_rows, updated_rows,
                   duplicate_strategy, started_at, completed_at,
                   rolled_back_at, created_at, created_by
            FROM import_jobs
            WHERE {where_clause}
            ORDER BY created_at DESC
            LIMIT :limit OFFSET :offset
        """
        rows = self.db.execute(text(query), params).fetchall()

        return {
            "total": total,
            "limit": limit,
            "offset": offset,
            "imports": [
                {
                    "import_id": r[0],
                    "status": r[1],
                    "filename": r[2],
                    "total_rows": r[3],
                    "imported": r[4] or 0,
                    "skipped": r[5] or 0,
                    "errors": r[6] or 0,
                    "updated": r[7] or 0,
                    "duplicate_strategy": r[8],
                    "started_at": r[9].isoformat() if r[9] else None,
                    "completed_at": r[10].isoformat() if r[10] else None,
                    "rolled_back_at": r[11].isoformat() if r[11] else None,
                    "created_at": r[12].isoformat() if r[12] else None,
                    "created_by": r[13],
                    "can_rollback": r[1] == "completed" and r[11] is None,
                }
                for r in rows
            ],
        }

    def get_import_status(self, import_id: str) -> Optional[Dict[str, Any]]:
        """Get status of a specific import, including in-flight progress.

        Checks in-memory progress first (for active imports), then falls
        back to the database record.
        """
        # Check in-flight progress
        progress = self._active_imports.get(import_id)
        if progress is not None:
            return {
                "import_id": import_id,
                "status": progress.status,
                "progress": {
                    "total_rows": progress.total_rows,
                    "processed_rows": progress.processed_rows,
                    "percentage": progress.percentage,
                },
                "results": {
                    "imported": progress.imported_rows,
                    "skipped": progress.skipped_rows,
                    "errors": progress.error_rows,
                    "updated": progress.updated_rows,
                },
                "is_active": True,
            }

        # Fall back to DB record
        row = self.db.execute(text("""
            SELECT id, status, filename, total_rows, processed_rows,
                   imported_rows, skipped_rows, error_rows, updated_rows,
                   duplicate_strategy, field_mapping, errors,
                   started_at, completed_at, rolled_back_at, created_at
            FROM import_jobs
            WHERE id = :id AND organization_id = :org_id
        """), {"id": import_id, "org_id": self.organization_id}).fetchone()

        if not row:
            return None

        total = row[3] or 1
        processed = row[4] or 0

        errors_data = row[11]
        if isinstance(errors_data, str):
            try:
                errors_data = json.loads(errors_data)
            except json.JSONDecodeError:
                errors_data = []

        mapping_data = row[10]
        if isinstance(mapping_data, str):
            try:
                mapping_data = json.loads(mapping_data)
            except json.JSONDecodeError:
                mapping_data = {}

        return {
            "import_id": row[0],
            "status": row[1],
            "filename": row[2],
            "progress": {
                "total_rows": row[3],
                "processed_rows": processed,
                "percentage": round((processed / total) * 100, 1) if total > 0 else 0,
            },
            "results": {
                "imported": row[5] or 0,
                "skipped": row[6] or 0,
                "errors": row[7] or 0,
                "updated": row[8] or 0,
            },
            "duplicate_strategy": row[9],
            "field_mapping": mapping_data if isinstance(mapping_data, dict) else {},
            "error_details": errors_data[:50] if isinstance(errors_data, list) else [],
            "timestamps": {
                "started_at": row[12].isoformat() if row[12] else None,
                "completed_at": row[13].isoformat() if row[13] else None,
                "rolled_back_at": row[14].isoformat() if row[14] else None,
                "created_at": row[15].isoformat() if row[15] else None,
            },
            "can_rollback": row[1] == "completed" and row[14] is None,
            "is_active": False,
        }

    # ------------------------------------------------------------------
    # Internal: core import execution
    # ------------------------------------------------------------------

    def _execute_import(
        self,
        row_generator: Generator,
        entity_type: str,
        field_mapping: Dict[str, str],
        filename: str,
        duplicate_strategy: str,
        default_stage: Optional[str],
        default_source: Optional[str],
        skip_invalid_rows: bool,
        validate_row_fn: Optional[Callable],
        transform_value_fn: Optional[Callable],
    ) -> ImportResult:
        """Core import logic shared between CSV and Excel.

        Uses savepoints for batch-level atomicity:
        - The entire import runs in one transaction
        - Each batch of BATCH_SIZE rows is wrapped in a savepoint
        - If skip_invalid_rows is False and a row fails validation,
          the current savepoint is rolled back and the import aborts
        - On success, a single commit writes everything
        """
        import_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)

        progress = ImportProgress(status="processing")
        self._active_imports[import_id] = progress

        try:
            return self._run_import(
                import_id=import_id,
                row_generator=row_generator,
                entity_type=entity_type,
                field_mapping=field_mapping,
                filename=filename,
                duplicate_strategy=duplicate_strategy,
                default_stage=default_stage,
                default_source=default_source,
                skip_invalid_rows=skip_invalid_rows,
                validate_row_fn=validate_row_fn,
                transform_value_fn=transform_value_fn,
                progress=progress,
                started_at=now,
            )
        finally:
            # Clean up in-flight tracker
            self._active_imports.pop(import_id, None)

    def _run_import(
        self,
        import_id: str,
        row_generator: Generator,
        entity_type: str,
        field_mapping: Dict[str, str],
        filename: str,
        duplicate_strategy: str,
        default_stage: Optional[str],
        default_source: Optional[str],
        skip_invalid_rows: bool,
        validate_row_fn: Optional[Callable],
        transform_value_fn: Optional[Callable],
        progress: ImportProgress,
        started_at: datetime,
    ) -> ImportResult:
        """Execute the import with savepoint-based batch processing."""
        # Ensure tracking tables exist
        self._ensure_tables()

        # Collect rows into batches for savepoint processing.
        # We buffer one batch at a time (BATCH_SIZE rows), not the whole file.
        batch: List[Tuple[Dict[str, Any], int]] = []  # (raw_row, row_num)
        headers_seen: Optional[List[str]] = None
        total_rows_counted = 0
        aborted = False
        abort_message = ""

        for headers, raw_row, row_num in row_generator:
            if headers_seen is None:
                headers_seen = headers
            total_rows_counted = row_num
            progress.total_rows = row_num

            # Apply field mapping
            mapped_row: Dict[str, Any] = {}
            for src_col, target_field in field_mapping.items():
                if src_col in raw_row:
                    value = raw_row[src_col]
                    if transform_value_fn:
                        value = transform_value_fn(target_field, value)
                    mapped_row[target_field] = value

            batch.append((mapped_row, row_num))

            if len(batch) >= BATCH_SIZE:
                abort_message = self._process_batch(
                    batch=batch,
                    entity_type=entity_type,
                    duplicate_strategy=duplicate_strategy,
                    default_stage=default_stage,
                    default_source=default_source,
                    skip_invalid_rows=skip_invalid_rows,
                    validate_row_fn=validate_row_fn,
                    progress=progress,
                )
                batch = []
                if abort_message:
                    aborted = True
                    break

        # Process remaining rows
        if batch and not aborted:
            abort_message = self._process_batch(
                batch=batch,
                entity_type=entity_type,
                duplicate_strategy=duplicate_strategy,
                default_stage=default_stage,
                default_source=default_source,
                skip_invalid_rows=skip_invalid_rows,
                validate_row_fn=validate_row_fn,
                progress=progress,
            )
            if abort_message:
                aborted = True

        # Finalize
        final_status = "failed" if aborted else "completed"
        now = datetime.now(timezone.utc)
        progress.status = final_status

        # Write the import job record (all in one commit with the data)
        self.db.execute(text("""
            INSERT INTO import_jobs
                (id, organization_id, created_by, status, filename,
                 total_rows, processed_rows, imported_rows, skipped_rows,
                 error_rows, updated_rows, duplicate_strategy, field_mapping,
                 errors, imported_lead_ids, started_at, completed_at, created_at)
            VALUES
                (:id, :org_id, :user_id, :status, :filename,
                 :total, :processed, :imported, :skipped,
                 :errors, :updated, :dup, :mapping,
                 :error_list, :lead_ids, :started, :completed, :created)
            ON CONFLICT (id) DO UPDATE SET
                status = EXCLUDED.status,
                processed_rows = EXCLUDED.processed_rows,
                imported_rows = EXCLUDED.imported_rows,
                skipped_rows = EXCLUDED.skipped_rows,
                error_rows = EXCLUDED.error_rows,
                updated_rows = EXCLUDED.updated_rows,
                errors = EXCLUDED.errors,
                imported_lead_ids = EXCLUDED.imported_lead_ids,
                completed_at = EXCLUDED.completed_at
        """), {
            "id": import_id,
            "org_id": self.organization_id,
            "user_id": self.user_id,
            "status": final_status,
            "filename": filename,
            "total": total_rows_counted,
            "processed": progress.processed_rows,
            "imported": progress.imported_rows,
            "skipped": progress.skipped_rows,
            "errors": progress.error_rows,
            "updated": progress.updated_rows,
            "dup": duplicate_strategy,
            "mapping": json.dumps(field_mapping),
            "error_list": json.dumps(progress.errors[:200]),
            "lead_ids": json.dumps(progress.imported_ids),
            "started": started_at,
            "completed": now,
            "created": started_at,
        })

        if aborted:
            # Roll back all data changes but keep the import_jobs record
            # We need a different approach: commit the job record first,
            # then the data is already rolled back by savepoint.
            # Actually — if we abort, we should NOT commit the data.
            # The savepoint rollback already undid the batch that failed.
            # Rows from earlier batches are still pending in the transaction.
            # We roll back the whole transaction and re-insert just the job record.
            self.db.rollback()

            # Re-insert the job record in a clean transaction
            self.db.execute(text("""
                INSERT INTO import_jobs
                    (id, organization_id, created_by, status, filename,
                     total_rows, processed_rows, imported_rows, skipped_rows,
                     error_rows, updated_rows, duplicate_strategy, field_mapping,
                     errors, imported_lead_ids, started_at, completed_at, created_at)
                VALUES
                    (:id, :org_id, :user_id, 'failed', :filename,
                     :total, :processed, 0, 0,
                     :errors, 0, :dup, :mapping,
                     :error_list, '[]', :started, :completed, :created)
                ON CONFLICT (id) DO UPDATE SET
                    status = 'failed',
                    errors = EXCLUDED.errors,
                    completed_at = EXCLUDED.completed_at
            """), {
                "id": import_id,
                "org_id": self.organization_id,
                "user_id": self.user_id,
                "filename": filename,
                "total": total_rows_counted,
                "processed": progress.processed_rows,
                "errors": progress.error_rows,
                "dup": duplicate_strategy,
                "mapping": json.dumps(field_mapping),
                "error_list": json.dumps(progress.errors[:200]),
                "started": started_at,
                "completed": now,
                "created": started_at,
            })
            self.db.commit()
        else:
            self.db.commit()

        message = abort_message if aborted else (
            f"Imported {progress.imported_rows} records"
            f"{f', updated {progress.updated_rows}' if progress.updated_rows else ''}"
            f"{f', skipped {progress.skipped_rows}' if progress.skipped_rows else ''}"
            f"{f', {progress.error_rows} errors' if progress.error_rows else ''}"
        )

        return ImportResult(
            import_id=import_id,
            status=final_status,
            filename=filename,
            total_rows=total_rows_counted,
            imported=progress.imported_rows,
            updated=progress.updated_rows,
            skipped=progress.skipped_rows,
            errors=progress.error_rows,
            error_details=progress.errors[:50],
            imported_ids=progress.imported_ids,
            field_mapping=field_mapping,
            message=message,
        )

    def _process_batch(
        self,
        batch: List[Tuple[Dict[str, Any], int]],
        entity_type: str,
        duplicate_strategy: str,
        default_stage: Optional[str],
        default_source: Optional[str],
        skip_invalid_rows: bool,
        validate_row_fn: Optional[Callable],
        progress: ImportProgress,
    ) -> str:
        """Process a batch of rows inside a savepoint.

        Returns empty string on success, or an abort message if
        skip_invalid_rows is False and validation fails.
        """
        savepoint = self.db.begin_nested()
        try:
            for mapped_row, row_num in batch:
                # Validate
                if validate_row_fn:
                    clean, row_errors = validate_row_fn(mapped_row, row_num)
                else:
                    clean = {k: v for k, v in mapped_row.items() if v is not None}
                    row_errors = []

                if row_errors:
                    progress.add_error(row_num, row_errors)
                    if not skip_invalid_rows:
                        savepoint.rollback()
                        return (
                            f"Aborted at row {row_num} due to validation error: "
                            f"{row_errors[0]}"
                        )
                    continue

                # Apply defaults
                if default_stage and ("stage" not in clean or not clean.get("stage")):
                    clean["stage"] = default_stage
                if default_source and ("source" not in clean or not clean.get("source")):
                    clean["source"] = default_source

                # Set organization scoping
                clean["organization_id"] = self.organization_id

                # Duplicate check (by email within org)
                existing_id = None
                if clean.get("email") and duplicate_strategy != "create":
                    query = f"""
                        SELECT id FROM {entity_type}
                        WHERE email = :email AND organization_id = :org_id
                        LIMIT 1
                    """
                    dup_row = self.db.execute(text(query), {
                        "email": clean["email"],
                        "org_id": self.organization_id,
                    }).fetchone()
                    if dup_row:
                        existing_id = dup_row[0]

                if existing_id and duplicate_strategy == "skip":
                    progress.add_skipped()
                    continue

                if existing_id and duplicate_strategy == "update":
                    update_fields = {
                        k: v for k, v in clean.items()
                        if v is not None and k != "organization_id"
                    }
                    if update_fields:
                        set_clauses = ", ".join(
                            f"{k} = :{k}" for k in update_fields
                        )
                        update_fields["_id"] = existing_id
                        query = f"""
                            UPDATE {entity_type}
                            SET {set_clauses}
                            WHERE id = :_id
                        """
                        self.db.execute(text(query), update_fields)
                    progress.add_updated(existing_id)
                    continue

                # Insert new record
                clean["created_at"] = datetime.now(timezone.utc)
                insert_cols = list(clean.keys())
                col_names = ", ".join(insert_cols)
                col_params = ", ".join(f":{c}" for c in insert_cols)

                query = f"""
                    INSERT INTO {entity_type} ({col_names})
                    VALUES ({col_params})
                    RETURNING id
                """
                result = self.db.execute(text(query), clean)
                new_id = result.fetchone()[0]
                progress.add_imported(new_id)

            savepoint.commit()
            return ""  # Success

        except Exception as e:
            logger.exception(f"Batch processing failed: {e}")
            savepoint.rollback()
            # Record error for each row in the batch that wasn't processed
            for mapped_row, row_num in batch:
                if progress.processed_rows < row_num:
                    progress.add_error(row_num, [f"Batch error: {str(e)}"])
            return f"Batch processing failed: {str(e)}"

    # ------------------------------------------------------------------
    # Internal: table bootstrap
    # ------------------------------------------------------------------

    def _ensure_tables(self) -> None:
        """Create import tracking tables if they don't exist.

        Uses IF NOT EXISTS so this is safe to call repeatedly.
        """
        self.db.execute(text("""
            CREATE TABLE IF NOT EXISTS import_jobs (
                id VARCHAR(36) PRIMARY KEY,
                organization_id INTEGER NOT NULL,
                created_by INTEGER NOT NULL,
                status VARCHAR(50) NOT NULL DEFAULT 'pending',
                filename VARCHAR(500),
                total_rows INTEGER DEFAULT 0,
                processed_rows INTEGER DEFAULT 0,
                imported_rows INTEGER DEFAULT 0,
                skipped_rows INTEGER DEFAULT 0,
                error_rows INTEGER DEFAULT 0,
                updated_rows INTEGER DEFAULT 0,
                duplicate_strategy VARCHAR(50) DEFAULT 'skip',
                field_mapping JSONB DEFAULT '{}',
                errors JSONB DEFAULT '[]',
                imported_lead_ids JSONB DEFAULT '[]',
                started_at TIMESTAMP WITH TIME ZONE,
                completed_at TIMESTAMP WITH TIME ZONE,
                rolled_back_at TIMESTAMP WITH TIME ZONE,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
            )
        """))
        self.db.execute(text("""
            CREATE TABLE IF NOT EXISTS import_field_mappings (
                id SERIAL PRIMARY KEY,
                organization_id INTEGER NOT NULL,
                name VARCHAR(255) NOT NULL,
                description TEXT,
                mappings JSONB NOT NULL DEFAULT '{}',
                created_by INTEGER,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
                updated_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
            )
        """))
        # Add index for history lookups
        self.db.execute(text("""
            CREATE INDEX IF NOT EXISTS idx_import_jobs_org_created
            ON import_jobs (organization_id, created_at DESC)
        """))
        self.db.flush()
