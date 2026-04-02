"""
Rate Sheet Parser Service
Handles parsing of uploaded rate sheets (PDF/Excel) using OCR and LLM.

Features:
- PDF text extraction using pdfplumber
- Excel file parsing using openpyxl
- LLM-based rate normalization and extraction
- Structured rate data output
"""

import os
import io
import json
import hashlib
import logging
from datetime import datetime, date, timezone
from typing import Dict, Any, List, Optional, Tuple
from decimal import Decimal
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError

logger = logging.getLogger(__name__)

# LLM Configuration
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")


class RateSheetParserService:
    """
    Service for parsing rate sheets from PDF and Excel files.
    Uses OCR for PDFs and structured parsing for Excel.
    Normalizes extracted data using LLM.
    """

    def __init__(self, db: Session):
        self.db = db
        self._anthropic_client = None
        self._openai_client = None

    def _get_anthropic_client(self):
        """Lazy load Anthropic client."""
        if self._anthropic_client is None and ANTHROPIC_API_KEY:
            try:
                import anthropic
                self._anthropic_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
            except Exception as e:
                logger.error(f"Failed to initialize Anthropic client: {e}")
        return self._anthropic_client

    def _get_openai_client(self):
        """Lazy load OpenAI client."""
        if self._openai_client is None and OPENAI_API_KEY:
            try:
                import openai
                self._openai_client = openai.OpenAI(api_key=OPENAI_API_KEY)
            except Exception as e:
                logger.error(f"Failed to initialize OpenAI client: {e}")
        return self._openai_client

    def calculate_file_hash(self, file_content: bytes) -> str:
        """Calculate SHA256 hash of file content."""
        return hashlib.sha256(file_content).hexdigest()

    async def parse_rate_sheet(
        self,
        file_content: bytes,
        filename: str,
        rate_sheet_id: int,
    ) -> Dict[str, Any]:
        """
        Parse a rate sheet file and extract rates.

        Args:
            file_content: Raw file bytes
            filename: Original filename
            rate_sheet_id: ID of the rate sheet record

        Returns:
            Dict with parsed rates and metadata
        """
        from models.rate_sheet import RateSheet, RateSheetRate

        # Update status to processing
        rate_sheet = self.db.query(RateSheet).filter(RateSheet.id == rate_sheet_id).first()
        if rate_sheet:
            rate_sheet.status = 'processing'
            self.db.commit()

        try:
            # Determine file type and extract text
            file_ext = filename.lower().split('.')[-1]

            if file_ext == 'pdf':
                extracted_text = self._extract_text_from_pdf(file_content)
            elif file_ext in ['xlsx', 'xls']:
                extracted_text = self._extract_text_from_excel(file_content)
            elif file_ext == 'csv':
                extracted_text = self._extract_text_from_csv(file_content)
            else:
                raise ValueError(f"Unsupported file type: {file_ext}")

            if not extracted_text or len(extracted_text.strip()) < 50:
                raise ValueError("Could not extract sufficient text from file")

            # Use LLM to parse the extracted text into structured rates
            parsed_result = await self._parse_with_llm(extracted_text, filename)

            if not parsed_result or 'rates' not in parsed_result:
                raise ValueError("LLM failed to parse rates from document")

            # Store parsed rates
            rates_created = []
            for rate_data in parsed_result.get('rates', []):
                rate_record = RateSheetRate(
                    rate_sheet_id=rate_sheet_id,
                    loan_type=rate_data.get('loan_type', 'conventional'),
                    loan_term=rate_data.get('loan_term', 30),
                    rate=Decimal(str(rate_data.get('rate', 0))),
                    apr=Decimal(str(rate_data.get('apr', 0))) if rate_data.get('apr') else None,
                    points=Decimal(str(rate_data.get('points', 0))) if rate_data.get('points') else None,
                    min_credit_score=rate_data.get('min_credit_score'),
                    max_credit_score=rate_data.get('max_credit_score'),
                    min_ltv=Decimal(str(rate_data.get('min_ltv', 0))) if rate_data.get('min_ltv') else None,
                    max_ltv=Decimal(str(rate_data.get('max_ltv', 0))) if rate_data.get('max_ltv') else None,
                    occupancy=rate_data.get('occupancy'),
                    property_type=rate_data.get('property_type'),
                    confidence_score=Decimal(str(rate_data.get('confidence', 0.8))),
                )
                self.db.add(rate_record)
                rates_created.append(rate_record)

            # Update rate sheet with parsed metadata
            if rate_sheet:
                rate_sheet.status = 'parsed'
                rate_sheet.parsed_data = {
                    'raw_text_length': len(extracted_text),
                    'rates_extracted': len(rates_created),
                    'parsing_model': parsed_result.get('model_used', 'unknown'),
                    'parsed_at': datetime.now(timezone.utc).isoformat(),
                }
                rate_sheet.lender_name = parsed_result.get('lender_name')
                if parsed_result.get('effective_date'):
                    try:
                        rate_sheet.effective_date = datetime.strptime(
                            parsed_result['effective_date'], '%Y-%m-%d'
                        ).date()
                    except (ValueError, TypeError):
                        pass

            self.db.commit()

            logger.info(f"Successfully parsed rate sheet {rate_sheet_id}: {len(rates_created)} rates extracted")

            return {
                'success': True,
                'rate_sheet_id': rate_sheet_id,
                'rates_count': len(rates_created),
                'lender_name': parsed_result.get('lender_name'),
                'effective_date': parsed_result.get('effective_date'),
                'rates': [r.to_dict() for r in rates_created],
            }

        except SQLAlchemyError as e:
            logger.error(f"Failed to parse rate sheet {rate_sheet_id}: {e}")

            # Update rate sheet with error
            if rate_sheet:
                rate_sheet.status = 'failed'
                rate_sheet.error_message = str(e)
                self.db.commit()

            return {
                'success': False,
                'rate_sheet_id': rate_sheet_id,
                'error': str(e),
            }

    def _extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF using pdfplumber."""
        try:
            import pdfplumber
        except ImportError:
            logger.error("pdfplumber not installed. Run: pip install pdfplumber")
            raise ImportError("pdfplumber required for PDF parsing")

        text_parts = []

        with pdfplumber.open(io.BytesIO(file_content)) as pdf:
            for page in pdf.pages:
                # Extract text
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)

                # Extract tables
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if row:
                            row_text = ' | '.join([str(cell) if cell else '' for cell in row])
                            text_parts.append(row_text)

        return '\n'.join(text_parts)

    def _extract_text_from_excel(self, file_content: bytes) -> str:
        """Extract text from Excel file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise ImportError("openpyxl required for Excel parsing")

        text_parts = []

        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")

            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                if row_values:
                    text_parts.append(' | '.join(row_values))

        return '\n'.join(text_parts)

    def _extract_text_from_csv(self, file_content: bytes) -> str:
        """Extract text from CSV file."""
        import csv

        text_parts = []
        content = file_content.decode('utf-8', errors='ignore')

        reader = csv.reader(io.StringIO(content))
        for row in reader:
            if row:
                text_parts.append(' | '.join(row))

        return '\n'.join(text_parts)

    async def _parse_with_llm(self, text: str, filename: str) -> Dict[str, Any]:
        """
        Use LLM to parse extracted text into structured rate data.
        """
        prompt = self._build_parsing_prompt(text, filename)

        # Try Anthropic first, then OpenAI
        anthropic_client = self._get_anthropic_client()
        if anthropic_client:
            return await self._parse_with_anthropic(anthropic_client, prompt)

        openai_client = self._get_openai_client()
        if openai_client:
            return await self._parse_with_openai(openai_client, prompt)

        # Fallback to simple pattern matching
        logger.warning("No LLM available, using fallback pattern matching")
        return self._fallback_pattern_parsing(text)

    def _build_parsing_prompt(self, text: str, filename: str) -> str:
        """Build the prompt for LLM rate extraction."""
        return f"""You are a mortgage rate sheet parser. Extract all interest rates from this rate sheet document.

DOCUMENT NAME: {filename}

DOCUMENT CONTENT:
{text[:15000]}  # Limit to prevent context overflow

TASK: Extract all mortgage rates from this document into a structured JSON format.

For each rate found, extract:
- loan_type: One of "conventional", "fha", "va", "usda", "jumbo", "non_qm"
- loan_term: Number of years (15, 20, 30, etc.)
- rate: Interest rate as a decimal (e.g., 6.875)
- apr: APR if available
- points: Discount points if available
- min_credit_score: Minimum credit score if specified
- max_credit_score: Maximum credit score if specified
- min_ltv: Minimum LTV if specified
- max_ltv: Maximum LTV if specified
- occupancy: "primary", "second_home", or "investment" if specified
- property_type: "single_family", "condo", "townhouse", "multi_family" if specified
- confidence: Your confidence in this extraction (0.0-1.0)

Also extract:
- lender_name: Name of the lender if visible
- effective_date: Date the rates are effective (format: YYYY-MM-DD)

RESPOND WITH VALID JSON ONLY in this exact format:
{{
    "lender_name": "Lender Name or null",
    "effective_date": "YYYY-MM-DD or null",
    "rates": [
        {{
            "loan_type": "conventional",
            "loan_term": 30,
            "rate": 6.875,
            "apr": 7.125,
            "points": 0,
            "min_credit_score": 720,
            "max_credit_score": null,
            "min_ltv": null,
            "max_ltv": 80,
            "occupancy": "primary",
            "property_type": null,
            "confidence": 0.95
        }}
    ]
}}

Extract ALL rates visible in the document. Focus on the most common rate scenarios."""

    async def _parse_with_anthropic(self, client, prompt: str) -> Dict[str, Any]:
        """Parse using Anthropic Claude."""
        try:
            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=4096,
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )

            response_text = message.content[0].text

            # Extract JSON from response
            result = self._extract_json_from_response(response_text)
            result['model_used'] = 'anthropic/claude-sonnet-4-20250514'
            return result

        except Exception as e:
            logger.error(f"Anthropic parsing failed: {e}")
            raise

    async def _parse_with_openai(self, client, prompt: str) -> Dict[str, Any]:
        """Parse using OpenAI GPT."""
        try:
            response = client.chat.completions.create(
                model="gpt-4-turbo-preview",
                messages=[
                    {"role": "user", "content": prompt}
                ],
                max_tokens=4096,
                temperature=0.1,
            )

            response_text = response.choices[0].message.content

            # Extract JSON from response
            result = self._extract_json_from_response(response_text)
            result['model_used'] = 'openai/gpt-4-turbo-preview'
            return result

        except Exception as e:
            logger.error(f"OpenAI parsing failed: {e}")
            raise

    def _extract_json_from_response(self, response_text: str) -> Dict[str, Any]:
        """Extract JSON object from LLM response."""
        # Try to find JSON in the response
        import re

        # Look for JSON block
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass

        # Try the whole response
        try:
            return json.loads(response_text)
        except json.JSONDecodeError:
            pass

        # Return empty result
        logger.warning("Could not extract JSON from LLM response")
        return {'rates': [], 'lender_name': None, 'effective_date': None}

    def _fallback_pattern_parsing(self, text: str) -> Dict[str, Any]:
        """
        Fallback pattern matching for rate extraction when no LLM available.
        Uses regex to find common rate patterns.
        """
        import re

        rates = []

        # Pattern for rates like "6.875%" or "6.875"
        rate_pattern = r'(\d+\.\d{2,3})%?'

        # Find all potential rates
        matches = re.findall(rate_pattern, text)

        for match in matches:
            rate_value = float(match)
            # Filter to reasonable mortgage rates (2-15%)
            if 2.0 <= rate_value <= 15.0:
                rates.append({
                    'loan_type': 'conventional',
                    'loan_term': 30,
                    'rate': rate_value,
                    'confidence': 0.5,  # Low confidence for pattern matching
                })

        # Remove duplicates
        seen_rates = set()
        unique_rates = []
        for rate in rates:
            if rate['rate'] not in seen_rates:
                seen_rates.add(rate['rate'])
                unique_rates.append(rate)

        return {
            'rates': unique_rates[:20],  # Limit to 20 rates
            'lender_name': None,
            'effective_date': None,
            'model_used': 'fallback_pattern_matching',
        }

    def get_best_rate_for_scenario(
        self,
        rate_sheet_id: int,
        loan_type: str = 'conventional',
        loan_term: int = 30,
        credit_score: Optional[int] = None,
        ltv: Optional[float] = None,
        occupancy: str = 'primary',
    ) -> Optional[Dict[str, Any]]:
        """
        Get the best matching rate from a rate sheet for a given scenario.
        """
        from models.rate_sheet import RateSheetRate

        query = self.db.query(RateSheetRate).filter(
            RateSheetRate.rate_sheet_id == rate_sheet_id,
            RateSheetRate.loan_type == loan_type,
            RateSheetRate.loan_term == loan_term,
        )

        # Filter by credit score if applicable
        if credit_score:
            query = query.filter(
                (RateSheetRate.min_credit_score.is_(None) | (RateSheetRate.min_credit_score <= credit_score)),
                (RateSheetRate.max_credit_score.is_(None) | (RateSheetRate.max_credit_score >= credit_score)),
            )

        # Filter by LTV if applicable
        if ltv:
            query = query.filter(
                (RateSheetRate.min_ltv.is_(None) | (RateSheetRate.min_ltv <= Decimal(str(ltv)))),
                (RateSheetRate.max_ltv.is_(None) | (RateSheetRate.max_ltv >= Decimal(str(ltv)))),
            )

        # Filter by occupancy if applicable
        query = query.filter(
            (RateSheetRate.occupancy.is_(None) | (RateSheetRate.occupancy == occupancy))
        )

        # Get the best (lowest) rate
        rate = query.order_by(RateSheetRate.rate.asc()).first()

        if rate:
            return rate.to_dict()
        return None
