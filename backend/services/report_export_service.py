"""
Report Export Service
=====================

Enterprise Readiness Domain 9: Analytics & Reporting

Provides:
1. PDF export for pipeline, scorecard, SLA compliance, and financial reports (Check 9.8)
2. Excel/XLSX export with formatted workbooks (Check 9.9)
3. SLA compliance report generation (Check 9.3)
4. CSV export (already exists, extended here for consistency)

Dependencies (already in requirements.txt):
    - reportlab >= 4.0.0
    - openpyxl >= 3.1.0
    - pandas >= 2.0.0

Usage:
    from services.report_export_service import report_exporter

    # Generate PDF
    pdf_bytes = report_exporter.generate_pdf(report_data, report_type="pipeline")

    # Generate Excel
    xlsx_bytes = report_exporter.generate_excel(report_data, report_type="scorecard")

    # Generate SLA compliance report
    sla_report = report_exporter.generate_sla_compliance_report(db, org_id=1)
"""

import io
import csv
import logging
from datetime import datetime, date, timedelta, timezone
from typing import Dict, Any, Optional, List
from decimal import Decimal

from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)


# =============================================================================
# PDF Generation
# =============================================================================

def generate_pdf(
    report_data: Dict[str, Any],
    report_type: str,
    title: Optional[str] = None,
    branding: Optional[Dict] = None,
) -> bytes:
    """
    Generate a PDF report from structured data.

    Args:
        report_data: The report data dict
        report_type: Type of report (pipeline, scorecard, sla_compliance, financial)
        title: Optional custom title
        branding: Optional tenant branding config (company_name, primary_color, logo_url)
                  from white_label_service.get_report_branding()

    Returns:
        PDF file as bytes
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

    # Use tenant branding colors or defaults
    primary_color = (branding or {}).get("primary_color", "#1a365d")
    company_name = (branding or {}).get("company_name", "The Tim Loss Team")

    buffer = io.BytesIO()
    page_size = landscape(letter) if report_type in ("pipeline", "sla_compliance") else letter
    doc = SimpleDocTemplate(buffer, pagesize=page_size, topMargin=0.75 * inch, bottomMargin=0.75 * inch)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Heading1"], fontSize=18, spaceAfter=12,
        textColor=colors.HexColor(primary_color),
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Heading2"], fontSize=12, spaceAfter=8,
        textColor=colors.HexColor("#4a5568"),
    )
    body_style = styles["Normal"]

    elements = []

    # Header with tenant branding
    report_title = title or _get_default_title(report_type)
    elements.append(Paragraph(report_title, title_style))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        subtitle_style,
    ))
    elements.append(Spacer(1, 12))

    # Build report-specific content
    if report_type == "pipeline":
        elements.extend(_build_pipeline_pdf(report_data, styles, subtitle_style))
    elif report_type == "scorecard":
        elements.extend(_build_scorecard_pdf(report_data, styles, subtitle_style))
    elif report_type == "sla_compliance":
        elements.extend(_build_sla_compliance_pdf(report_data, styles, subtitle_style))
    elif report_type == "financial":
        elements.extend(_build_financial_pdf(report_data, styles, subtitle_style))
    else:
        # Generic table export
        elements.extend(_build_generic_pdf(report_data, styles, subtitle_style))

    # Footer with timestamp
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"{company_name} — Confidential | {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}",
        ParagraphStyle("Footer", parent=body_style, fontSize=8, textColor=colors.gray),
    ))

    doc.build(elements)
    return buffer.getvalue()


def _get_default_title(report_type: str) -> str:
    titles = {
        "pipeline": "Pipeline Report",
        "scorecard": "Loan Officer Scorecard",
        "sla_compliance": "SLA Compliance Report",
        "financial": "Financial Report",
    }
    return titles.get(report_type, "Report")


def _make_table(headers: List[str], rows: List[List], col_widths: Optional[List] = None):
    """Create a styled table."""
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle

    data = [headers] + rows
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b6cb0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])
    return table


def _build_pipeline_pdf(data, styles, subtitle_style):
    from reportlab.platypus import Paragraph, Spacer

    elements = []

    # Summary metrics
    if "metrics" in data:
        m = data["metrics"]
        elements.append(Paragraph("Pipeline Summary", subtitle_style))
        summary_rows = [
            ["Total Loans", str(m.get("total_count", 0))],
            ["Total Volume", m.get("total_volume_formatted", "$0")],
            ["Closing Soon", str(m.get("closing_soon", 0))],
            ["Avg Days in Status", str(m.get("avg_days_in_status", 0))],
        ]
        elements.append(_make_table(["Metric", "Value"], summary_rows))
        elements.append(Spacer(1, 12))

    # Loans table
    if "loans" in data:
        elements.append(Paragraph("Loan Details", subtitle_style))
        headers = ["Loan #", "Borrower", "Amount", "Status", "Days", "Close Date"]
        rows = []
        for loan in data["loans"][:100]:
            rows.append([
                str(loan.get("loan_number", "")),
                str(loan.get("borrower", "")),
                str(loan.get("amount_formatted", "")),
                str(loan.get("status", "")),
                str(loan.get("days_in_status", "")),
                str(loan.get("expected_close", "")),
            ])
        elements.append(_make_table(headers, rows))

    return elements


def _build_scorecard_pdf(data, styles, subtitle_style):
    from reportlab.platypus import Paragraph, Spacer

    elements = []
    elements.append(Paragraph("Scorecard Metrics", subtitle_style))

    if "conversion" in data:
        c = data["conversion"]
        rows = [
            ["Lead to Application", f"{c.get('lead_to_app', 0)}%"],
            ["Application to Submission", f"{c.get('app_to_submit', 0)}%"],
            ["Submission to Approval", f"{c.get('submit_to_approve', 0)}%"],
            ["Overall Pull-Through", f"{c.get('overall_pull_through', 0)}%"],
        ]
        elements.append(_make_table(["Stage", "Rate"], rows))
        elements.append(Spacer(1, 12))

    if "funded" in data:
        f = data["funded"]
        rows = [
            ["Funded Units", str(f.get("count", 0))],
            ["Funded Volume", f.get("volume_formatted", "$0")],
            ["Avg Loan Size", f.get("avg_size_formatted", "$0")],
        ]
        elements.append(Paragraph("Funding Summary", subtitle_style))
        elements.append(_make_table(["Metric", "Value"], rows))

    return elements


def _build_sla_compliance_pdf(data, styles, subtitle_style):
    from reportlab.platypus import Paragraph, Spacer

    elements = []

    # Overall compliance
    if "overall" in data:
        o = data["overall"]
        elements.append(Paragraph("Overall SLA Compliance", subtitle_style))
        rows = [
            ["Compliance Rate", f"{o.get('compliance_rate', 0):.1f}%"],
            ["Total Milestones", str(o.get("total_milestones", 0))],
            ["On Time", str(o.get("on_time", 0))],
            ["Late", str(o.get("late", 0))],
            ["Overdue", str(o.get("overdue", 0))],
        ]
        elements.append(_make_table(["Metric", "Value"], rows))
        elements.append(Spacer(1, 12))

    # By milestone type
    if "by_milestone" in data:
        elements.append(Paragraph("Compliance by Milestone", subtitle_style))
        headers = ["Milestone", "Total", "On Time", "Late", "Compliance %", "Avg Hours"]
        rows = []
        for m in data["by_milestone"]:
            rows.append([
                str(m.get("milestone_type", "")),
                str(m.get("total", 0)),
                str(m.get("on_time", 0)),
                str(m.get("late", 0)),
                f"{m.get('compliance_rate', 0):.1f}%",
                f"{m.get('avg_completion_hours', 0):.1f}",
            ])
        elements.append(_make_table(headers, rows))
        elements.append(Spacer(1, 12))

    # Bottlenecks
    if "bottlenecks" in data and data["bottlenecks"]:
        elements.append(Paragraph("Bottleneck Analysis", subtitle_style))
        headers = ["Stage", "Avg Delay (hours)", "Affected Loans"]
        rows = [[str(b.get("stage", "")), f"{b.get('avg_delay_hours', 0):.1f}", str(b.get("count", 0))]
                for b in data["bottlenecks"]]
        elements.append(_make_table(headers, rows))

    return elements


def _build_financial_pdf(data, styles, subtitle_style):
    from reportlab.platypus import Paragraph, Spacer

    elements = []
    elements.append(Paragraph("Financial Summary", subtitle_style))

    if "sections" in data:
        for section in data["sections"]:
            elements.append(Paragraph(section.get("name", ""), subtitle_style))
            if "items" in section:
                headers = ["Account", "Amount"]
                rows = [[str(item.get("name", "")), str(item.get("amount_formatted", ""))]
                        for item in section["items"]]
                elements.append(_make_table(headers, rows))
                elements.append(Spacer(1, 8))

    return elements


def _build_generic_pdf(data, styles, subtitle_style):
    from reportlab.platypus import Paragraph, Spacer

    elements = []

    # Try to render any list of dicts as a table
    for key, value in data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            elements.append(Paragraph(key.replace("_", " ").title(), subtitle_style))
            headers = list(value[0].keys())
            rows = [[str(item.get(h, "")) for h in headers] for item in value[:100]]
            elements.append(_make_table(headers, rows))
            elements.append(Spacer(1, 12))

    return elements


# =============================================================================
# Excel Generation
# =============================================================================

def generate_excel(report_data: Dict[str, Any], report_type: str, title: Optional[str] = None) -> bytes:
    """
    Generate an Excel workbook from structured data.

    Args:
        report_data: The report data dict
        report_type: Type of report
        title: Optional custom title

    Returns:
        XLSX file as bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="1A365D")
    subtitle_font = Font(name="Calibri", bold=True, size=11, color="4A5568")
    currency_format = '#,##0.00'
    pct_format = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def auto_width(ws, min_width=10, max_width=40):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))

    report_title = title or _get_default_title(report_type)

    if report_type == "pipeline":
        _build_pipeline_excel(wb, report_data, report_title, header_font, header_fill,
                              title_font, subtitle_font, thin_border, style_header_row, auto_width)
    elif report_type == "scorecard":
        _build_scorecard_excel(wb, report_data, report_title, header_font, header_fill,
                               title_font, subtitle_font, thin_border, style_header_row, auto_width)
    elif report_type == "sla_compliance":
        _build_sla_compliance_excel(wb, report_data, report_title, header_font, header_fill,
                                     title_font, subtitle_font, thin_border, style_header_row, auto_width)
    else:
        _build_generic_excel(wb, report_data, report_title, header_font, header_fill,
                             title_font, thin_border, style_header_row, auto_width)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_pipeline_excel(wb, data, title, header_font, header_fill,
                          title_font, subtitle_font, border, style_header, auto_width):
    ws = wb.active
    ws.title = "Pipeline"

    # Title
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = subtitle_font

    # Summary
    if "metrics" in data:
        m = data["metrics"]
        ws.cell(row=4, column=1, value="Pipeline Summary").font = subtitle_font
        for i, (label, val) in enumerate([
            ("Total Loans", m.get("total_count", 0)),
            ("Total Volume", m.get("total_volume_formatted", "$0")),
            ("Closing Soon", m.get("closing_soon", 0)),
        ]):
            ws.cell(row=5 + i, column=1, value=label)
            ws.cell(row=5 + i, column=2, value=val)

    # Loan details
    if "loans" in data:
        row = 10
        ws.cell(row=row, column=1, value="Loan Details").font = subtitle_font
        row += 1
        headers = ["Loan #", "Borrower", "Amount", "Type", "Status", "Days in Status", "Close Date", "LO"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header(ws, row, len(headers))

        for loan in data["loans"][:500]:
            row += 1
            ws.cell(row=row, column=1, value=loan.get("loan_number", ""))
            ws.cell(row=row, column=2, value=loan.get("borrower", ""))
            ws.cell(row=row, column=3, value=loan.get("amount", 0))
            ws.cell(row=row, column=4, value=loan.get("loan_type", ""))
            ws.cell(row=row, column=5, value=loan.get("status", ""))
            ws.cell(row=row, column=6, value=loan.get("days_in_status", 0))
            ws.cell(row=row, column=7, value=loan.get("expected_close", ""))
            ws.cell(row=row, column=8, value=loan.get("lo_name", ""))

    auto_width(ws)


def _build_scorecard_excel(wb, data, title, header_font, header_fill,
                            title_font, subtitle_font, border, style_header, auto_width):
    ws = wb.active
    ws.title = "Scorecard"
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = subtitle_font

    row = 4
    if "conversion" in data:
        ws.cell(row=row, column=1, value="Conversion Rates").font = subtitle_font
        row += 1
        for label, key in [
            ("Lead to App", "lead_to_app"), ("App to Submit", "app_to_submit"),
            ("Submit to Approve", "submit_to_approve"), ("Pull-Through", "overall_pull_through"),
        ]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=f"{data['conversion'].get(key, 0)}%")
            row += 1

    if "funded" in data:
        row += 1
        ws.cell(row=row, column=1, value="Funding Summary").font = subtitle_font
        row += 1
        for label, key in [("Units", "count"), ("Volume", "volume_formatted"), ("Avg Size", "avg_size_formatted")]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=data["funded"].get(key, ""))
            row += 1

    auto_width(ws)


def _build_sla_compliance_excel(wb, data, title, header_font, header_fill,
                                 title_font, subtitle_font, border, style_header, auto_width):
    ws = wb.active
    ws.title = "SLA Compliance"
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = subtitle_font

    # Overall
    row = 4
    if "overall" in data:
        o = data["overall"]
        ws.cell(row=row, column=1, value="Overall Compliance").font = subtitle_font
        row += 1
        for label, val in [
            ("Compliance Rate", f"{o.get('compliance_rate', 0):.1f}%"),
            ("Total Milestones", o.get("total_milestones", 0)),
            ("On Time", o.get("on_time", 0)),
            ("Late", o.get("late", 0)),
            ("Overdue", o.get("overdue", 0)),
        ]:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val)
            row += 1

    # By milestone
    if "by_milestone" in data:
        row += 1
        ws.cell(row=row, column=1, value="By Milestone Type").font = subtitle_font
        row += 1
        headers = ["Milestone", "Total", "On Time", "Late", "Compliance %", "Avg Hours"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        style_header(ws, row, len(headers))

        for m in data["by_milestone"]:
            row += 1
            ws.cell(row=row, column=1, value=m.get("milestone_type", ""))
            ws.cell(row=row, column=2, value=m.get("total", 0))
            ws.cell(row=row, column=3, value=m.get("on_time", 0))
            ws.cell(row=row, column=4, value=m.get("late", 0))
            ws.cell(row=row, column=5, value=f"{m.get('compliance_rate', 0):.1f}%")
            ws.cell(row=row, column=6, value=round(m.get("avg_completion_hours", 0), 1))

    # Bottlenecks sheet
    if "bottlenecks" in data and data["bottlenecks"]:
        ws2 = wb.create_sheet("Bottlenecks")
        ws2.cell(row=1, column=1, value="Bottleneck Analysis").font = title_font
        headers = ["Stage", "Avg Delay (hours)", "Affected Loans"]
        for col, h in enumerate(headers, 1):
            ws2.cell(row=3, column=col, value=h)
        style_header(ws2, 3, len(headers))
        for i, b in enumerate(data["bottlenecks"]):
            ws2.cell(row=4 + i, column=1, value=b.get("stage", ""))
            ws2.cell(row=4 + i, column=2, value=round(b.get("avg_delay_hours", 0), 1))
            ws2.cell(row=4 + i, column=3, value=b.get("count", 0))
        auto_width(ws2)

    auto_width(ws)


def _build_generic_excel(wb, data, title, header_font, header_fill,
                         title_font, border, style_header, auto_width):
    ws = wb.active
    ws.title = "Report"
    ws.cell(row=1, column=1, value=title).font = title_font

    row = 3
    for key, value in data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = title_font
            row += 1
            headers = list(value[0].keys())
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            style_header(ws, row, len(headers))
            for item in value[:1000]:
                row += 1
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=str(item.get(h, "")))
            row += 2

    auto_width(ws)


# =============================================================================
# SLA Compliance Report Generation (Check 9.3)
# =============================================================================

def generate_sla_compliance_report(
    db,
    org_id: int,
    period_days: int = 30,
    scope_type: str = "organization",
    scope_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Generate comprehensive SLA compliance report from database.

    Args:
        db: SQLAlchemy session
        org_id: Organization ID
        period_days: Lookback period in days
        scope_type: "organization", "team", or "user"
        scope_id: Optional team_id or user_id for scoping

    Returns:
        Report data dict ready for PDF/Excel/JSON export
    """
    from sqlalchemy import text

    params = {"org_id": org_id, "period_days": period_days}
    scope_filter = "lmh.organization_id = :org_id"
    if scope_type == "user" and scope_id:
        scope_filter += " AND lmh.assigned_to_id = :scope_id"
        params["scope_id"] = scope_id
    elif scope_type == "team" and scope_id:
        scope_filter += " AND lmh.team_id = :scope_id"
        params["scope_id"] = scope_id

    try:
        # Overall compliance
        overall_sql = """
            SELECT
                COUNT(*) AS total_milestones,
                COUNT(CASE WHEN status = 'COMPLETED' AND variance_hours <= 0 THEN 1 END) AS on_time,
                COUNT(CASE WHEN status = 'COMPLETED' AND variance_hours > 0 THEN 1 END) AS late,
                COUNT(CASE WHEN status = 'OVERDUE' THEN 1 END) AS overdue,
                COUNT(CASE WHEN status IN ('IN_PROGRESS', 'ON_TRACK') THEN 1 END) AS in_progress,
                AVG(CASE WHEN status = 'COMPLETED' THEN actual_hours END) AS avg_completion_hours,
                AVG(CASE WHEN status = 'COMPLETED' THEN variance_pct END) AS avg_variance_pct
            FROM loan_milestone_history lmh
            WHERE """ + scope_filter + """
                AND lmh.created_at >= CURRENT_DATE - :period_days
        """
        overall = db.execute(text(overall_sql), params).fetchone()

        total = overall[0] or 0
        on_time = overall[1] or 0
        late = overall[2] or 0
        overdue = overall[3] or 0
        compliance_rate = (on_time / total * 100) if total > 0 else 0

        # By milestone type
        milestone_sql = """
            SELECT
                milestone_type,
                COUNT(*) AS total,
                COUNT(CASE WHEN status = 'COMPLETED' AND variance_hours <= 0 THEN 1 END) AS on_time,
                COUNT(CASE WHEN status = 'COMPLETED' AND variance_hours > 0 THEN 1 END) AS late,
                AVG(CASE WHEN status = 'COMPLETED' THEN actual_hours END) AS avg_completion_hours
            FROM loan_milestone_history lmh
            WHERE """ + scope_filter + """
                AND lmh.created_at >= CURRENT_DATE - :period_days
            GROUP BY milestone_type
            ORDER BY COUNT(*) DESC
        """
        by_milestone_rows = db.execute(text(milestone_sql), params).fetchall()

        by_milestone = []
        for row in by_milestone_rows:
            mt_total = row[1] or 0
            mt_on_time = row[2] or 0
            by_milestone.append({
                "milestone_type": str(row[0]),
                "total": mt_total,
                "on_time": mt_on_time,
                "late": row[3] or 0,
                "compliance_rate": (mt_on_time / mt_total * 100) if mt_total > 0 else 0,
                "avg_completion_hours": float(row[4] or 0),
            })

        # Bottlenecks (stages with highest average delay)
        bottleneck_sql = """
            SELECT
                milestone_type AS stage,
                AVG(variance_hours) AS avg_delay_hours,
                COUNT(*) AS count
            FROM loan_milestone_history lmh
            WHERE """ + scope_filter + """
                AND lmh.created_at >= CURRENT_DATE - :period_days
                AND variance_hours > 0
                AND status = 'COMPLETED'
            GROUP BY milestone_type
            HAVING AVG(variance_hours) > 2
            ORDER BY AVG(variance_hours) DESC
            LIMIT 10
        """
        bottleneck_rows = db.execute(text(bottleneck_sql), params).fetchall()

        bottlenecks = [
            {"stage": str(row[0]), "avg_delay_hours": float(row[1] or 0), "count": row[2] or 0}
            for row in bottleneck_rows
        ]

        # Alerts summary
        alerts_summary = db.execute(text("""
            SELECT
                alert_type,
                COUNT(*) AS count,
                COUNT(CASE WHEN status = 'ACTIVE' THEN 1 END) AS active
            FROM sla_alerts
            WHERE organization_id = :org_id
                AND triggered_at >= CURRENT_DATE - :period_days
            GROUP BY alert_type
        """), params).fetchall()

        alerts = [
            {"type": row[0], "total": row[1], "active": row[2]}
            for row in alerts_summary
        ]

        return {
            "report_type": "sla_compliance",
            "organization_id": org_id,
            "period_days": period_days,
            "generated_at": datetime.now(timezone.utc).isoformat() + "Z",
            "scope": {"type": scope_type, "id": scope_id},
            "overall": {
                "compliance_rate": round(compliance_rate, 1),
                "total_milestones": total,
                "on_time": on_time,
                "late": late,
                "overdue": overdue,
                "in_progress": overall[4] or 0,
                "avg_completion_hours": round(float(overall[5] or 0), 1),
                "avg_variance_pct": round(float(overall[6] or 0), 1),
            },
            "by_milestone": by_milestone,
            "bottlenecks": bottlenecks,
            "alerts": alerts,
        }

    except Exception as e:
        logger.error(f"SLA compliance report generation failed: {e}")
        return {
            "report_type": "sla_compliance",
            "organization_id": org_id,
            "period_days": period_days,
            "generated_at": datetime.now(timezone.utc).isoformat() + "Z",
            "error": str(e),
            "overall": {
                "compliance_rate": 0, "total_milestones": 0,
                "on_time": 0, "late": 0, "overdue": 0, "in_progress": 0,
                "avg_completion_hours": 0, "avg_variance_pct": 0,
            },
            "by_milestone": [],
            "bottlenecks": [],
            "alerts": [],
        }


# =============================================================================
# CSV Export (Extended for consistency)
# =============================================================================

def generate_csv(report_data: Dict[str, Any], report_type: str) -> bytes:
    """Generate CSV from report data."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Find the primary list of dicts in the data
    rows_key = None
    for key, value in report_data.items():
        if isinstance(value, list) and value and isinstance(value[0], dict):
            rows_key = key
            break

    if rows_key:
        items = report_data[rows_key]
        headers = list(items[0].keys())
        writer.writerow(headers)
        for item in items:
            writer.writerow([str(item.get(h, "")) for h in headers])
    else:
        # Flat key-value export
        writer.writerow(["Key", "Value"])
        for key, value in report_data.items():
            if not isinstance(value, (list, dict)):
                writer.writerow([key, str(value)])

    return buffer.getvalue().encode("utf-8")


# =============================================================================
# General-Purpose Export Helpers (StreamingResponse wrappers)
# =============================================================================

def export_to_csv(
    data: List[Dict[str, Any]],
    columns: List[str],
    filename: str = "export.csv",
) -> StreamingResponse:
    """
    Export a list of dicts to CSV with proper escaping and UTF-8 BOM.

    Args:
        data: Rows as list of dicts.
        columns: Ordered list of column keys to include.
        filename: Download filename.

    Returns:
        FastAPI StreamingResponse ready to return from an endpoint.
    """
    from fastapi.responses import StreamingResponse

    buffer = io.BytesIO()
    # UTF-8 BOM so Excel opens with correct encoding
    buffer.write(b"\xef\xbb\xbf")

    text_wrapper = io.TextIOWrapper(buffer, encoding="utf-8", newline="", write_through=True)
    writer = csv.DictWriter(text_wrapper, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        # Coerce values to string-safe representations
        safe_row = {}
        for col in columns:
            val = row.get(col, "")
            if val is None:
                safe_row[col] = ""
            elif isinstance(val, Decimal):
                safe_row[col] = str(val)
            elif isinstance(val, (datetime, date)):
                safe_row[col] = val.isoformat()
            else:
                safe_row[col] = val
        writer.writerow(safe_row)

    text_wrapper.detach()  # release without closing underlying buffer
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_to_excel(
    data: List[Dict[str, Any]],
    columns: List[str],
    sheet_name: str = "Report",
    filename: str = "export.xlsx",
) -> StreamingResponse:
    """
    Export a list of dicts to an XLSX workbook using openpyxl.

    Args:
        data: Rows as list of dicts.
        columns: Ordered list of column keys to include.
        sheet_name: Worksheet name.
        filename: Download filename.

    Returns:
        FastAPI StreamingResponse ready to return from an endpoint.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # Write header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            if isinstance(val, Decimal):
                val = float(val)
            elif isinstance(val, (datetime, date)):
                val = val.isoformat()
            elif val is None:
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val).border = thin_border

    # Auto-fit column widths
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(len(data) + 2, 102)):  # sample first 100 rows
            cell_len = len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =============================================================================
# Report Formatters (query DB and return flat list[dict] for export)
# =============================================================================

def format_pipeline_report(
    db,
    org_id: int,
    scope: Dict[str, Any],
    filters: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """
    Build a flat pipeline report from leads and loans.

    Args:
        db: SQLAlchemy session.
        org_id: Organization ID.
        scope: Scope dict from get_report_scope().
        filters: Optional dict with keys like 'stage', 'date_from', 'date_to'.

    Returns:
        List of dicts ready for CSV/Excel export.
    """
    from database.models import Lead, Loan, User

    filters = filters or {}

    # --- Leads ---
    from middleware.report_access import apply_scope_to_lead_query, apply_scope_to_loan_query

    lead_q = db.query(
        Lead.id,
        Lead.name,
        Lead.email,
        Lead.phone,
        Lead.stage,
        Lead.source,
        Lead.ai_score,
        Lead.loan_amount,
        Lead.loan_type,
        Lead.created_at,
    )
    lead_q = apply_scope_to_lead_query(lead_q, scope, Lead)

    if filters.get("stage"):
        lead_q = lead_q.filter(Lead.stage == filters["stage"])
    if filters.get("date_from"):
        lead_q = lead_q.filter(Lead.created_at >= filters["date_from"])
    if filters.get("date_to"):
        lead_q = lead_q.filter(Lead.created_at <= filters["date_to"])

    lead_q = lead_q.order_by(Lead.created_at.desc()).limit(5000)
    leads = lead_q.all()

    # --- Loans ---
    loan_q = db.query(
        Loan.id,
        Loan.loan_number,
        Loan.borrower_name,
        Loan.stage,
        Loan.loan_type,
        Loan.amount,
        Loan.rate,
        Loan.days_in_stage,
        Loan.closing_date,
        Loan.funded_date,
        Loan.created_at,
        Loan.loan_officer_name,
    )
    loan_q = apply_scope_to_loan_query(loan_q, scope, Loan)

    if filters.get("stage"):
        loan_q = loan_q.filter(Loan.stage == filters["stage"])
    if filters.get("date_from"):
        loan_q = loan_q.filter(Loan.created_at >= filters["date_from"])
    if filters.get("date_to"):
        loan_q = loan_q.filter(Loan.created_at <= filters["date_to"])

    loan_q = loan_q.order_by(Loan.created_at.desc()).limit(5000)
    loans = loan_q.all()

    rows: List[Dict[str, Any]] = []

    for ld in leads:
        rows.append({
            "record_type": "Lead",
            "id": ld.id,
            "name": ld.name or "",
            "email": ld.email or "",
            "phone": ld.phone or "",
            "stage": ld.stage or "",
            "source": ld.source or "",
            "ai_score": ld.ai_score,
            "loan_type": ld.loan_type or "",
            "amount": ld.loan_amount,
            "rate": None,
            "days_in_stage": None,
            "loan_number": None,
            "loan_officer": None,
            "closing_date": None,
            "funded_date": None,
            "created_at": ld.created_at,
        })

    for ln in loans:
        rows.append({
            "record_type": "Loan",
            "id": ln.id,
            "name": ln.borrower_name or "",
            "email": "",
            "phone": "",
            "stage": ln.stage or "",
            "source": "",
            "ai_score": None,
            "loan_type": ln.loan_type or "",
            "amount": ln.amount,
            "rate": ln.rate,
            "days_in_stage": ln.days_in_stage,
            "loan_number": ln.loan_number or "",
            "loan_officer": ln.loan_officer_name or "",
            "closing_date": ln.closing_date,
            "funded_date": ln.funded_date,
            "created_at": ln.created_at,
        })

    return rows


PIPELINE_COLUMNS = [
    "record_type", "id", "name", "email", "phone", "stage", "source",
    "ai_score", "loan_type", "amount", "rate", "days_in_stage",
    "loan_number", "loan_officer", "closing_date", "funded_date", "created_at",
]


def format_production_report(
    db,
    org_id: int,
    scope: Dict[str, Any],
    date_from: Optional["date"] = None,
    date_to: Optional["date"] = None,
) -> List[Dict[str, Any]]:
    """
    Build an LO production scorecard.

    Each row represents one user with their key production metrics.

    Args:
        db: SQLAlchemy session.
        org_id: Organization ID.
        scope: Scope dict from get_report_scope().
        date_from: Start of date range (defaults to 30 days ago).
        date_to: End of date range (defaults to today).

    Returns:
        List of dicts ready for CSV/Excel export.
    """
    from sqlalchemy import func, case
    from database.models import User, Loan, Lead
    from database.enums import LoanStage

    if date_from is None:
        date_from = date.today() - timedelta(days=30)
    if date_to is None:
        date_to = date.today()

    # Determine which users to include
    user_q = db.query(User).filter(
        User.organization_id == org_id,
        User.is_active == True,
    )
    if scope["scope"] == "user":
        user_q = user_q.filter(User.id == scope.get("user_id"))
    elif scope["scope"] == "branch":
        branch_id = scope.get("branch_id")
        if branch_id:
            user_q = user_q.filter(User.branch_id == branch_id)

    users = user_q.all()
    user_ids = [u.id for u in users]
    user_map = {u.id: u for u in users}

    if not user_ids:
        return []

    # Aggregate loan metrics per LO
    loan_agg = db.query(
        Loan.loan_officer_id,
        func.count(Loan.id).label("total_loans"),
        func.count(case((Loan.stage == LoanStage.FUNDED, 1))).label("funded_count"),
        func.sum(case((Loan.stage == LoanStage.FUNDED, Loan.amount), else_=0)).label("funded_volume"),
        func.sum(Loan.amount).label("total_pipeline_volume"),
        func.avg(Loan.days_in_stage).label("avg_days_in_stage"),
    ).filter(
        Loan.organization_id == org_id,
        Loan.loan_officer_id.in_(user_ids),
        Loan.created_at >= date_from,
        Loan.created_at <= date_to + timedelta(days=1),
    ).group_by(Loan.loan_officer_id).all()

    loan_map = {row.loan_officer_id: row for row in loan_agg}

    # Aggregate lead counts per LO
    lead_agg = db.query(
        Lead.owner_id,
        func.count(Lead.id).label("total_leads"),
    ).filter(
        Lead.organization_id == org_id,
        Lead.owner_id.in_(user_ids),
        Lead.created_at >= date_from,
        Lead.created_at <= date_to + timedelta(days=1),
    ).group_by(Lead.owner_id).all()

    lead_map = {row.owner_id: row.total_leads for row in lead_agg}

    rows: List[Dict[str, Any]] = []
    for uid in user_ids:
        u = user_map[uid]
        lm = loan_map.get(uid)
        total_loans = lm.total_loans if lm else 0
        funded = lm.funded_count if lm else 0
        funded_vol = float(lm.funded_volume or 0) if lm else 0.0
        pipeline_vol = float(lm.total_pipeline_volume or 0) if lm else 0.0
        avg_days = round(float(lm.avg_days_in_stage or 0), 1) if lm else 0.0
        total_leads = lead_map.get(uid, 0)
        pull_through = round((funded / total_loans * 100), 1) if total_loans > 0 else 0.0

        rows.append({
            "user_id": uid,
            "name": u.full_name,
            "email": u.email,
            "role": u.permission_role or u.role or "",
            "total_leads": total_leads,
            "total_loans": total_loans,
            "funded_count": funded,
            "funded_volume": funded_vol,
            "pipeline_volume": pipeline_vol,
            "pull_through_pct": pull_through,
            "avg_days_in_stage": avg_days,
            "period_from": str(date_from),
            "period_to": str(date_to),
        })

    # Sort by funded volume descending
    rows.sort(key=lambda r: r["funded_volume"], reverse=True)
    return rows


PRODUCTION_COLUMNS = [
    "name", "email", "role", "total_leads", "total_loans",
    "funded_count", "funded_volume", "pipeline_volume",
    "pull_through_pct", "avg_days_in_stage", "period_from", "period_to",
]


def format_compliance_report(
    db,
    org_id: int,
    scope: Dict[str, Any],
    period_days: int = 30,
) -> List[Dict[str, Any]]:
    """
    Build a compliance export: one row per loan with SLA and compliance flags.

    Args:
        db: SQLAlchemy session.
        org_id: Organization ID.
        scope: Scope dict from get_report_scope().
        period_days: Lookback window in days.

    Returns:
        List of dicts ready for CSV/Excel export.
    """
    from database.models import Loan
    from middleware.report_access import apply_scope_to_loan_query

    cutoff = datetime.now(timezone.utc) - timedelta(days=period_days)

    loan_q = db.query(
        Loan.id,
        Loan.loan_number,
        Loan.borrower_name,
        Loan.stage,
        Loan.loan_type,
        Loan.amount,
        Loan.days_in_stage,
        Loan.sla_status,
        Loan.risk_score,
        Loan.closing_date,
        Loan.funded_date,
        Loan.lock_expiration_date,
        Loan.loan_officer_name,
        Loan.created_at,
    ).filter(
        Loan.created_at >= cutoff,
    )
    loan_q = apply_scope_to_loan_query(loan_q, scope, Loan)
    loan_q = loan_q.order_by(Loan.created_at.desc()).limit(5000)
    loans = loan_q.all()

    rows: List[Dict[str, Any]] = []
    now = datetime.now(timezone.utc)
    for ln in loans:
        lock_expired = False
        if ln.lock_expiration_date and ln.lock_expiration_date < now:
            lock_expired = True

        rows.append({
            "loan_id": ln.id,
            "loan_number": ln.loan_number or "",
            "borrower": ln.borrower_name or "",
            "stage": ln.stage or "",
            "loan_type": ln.loan_type or "",
            "amount": ln.amount,
            "days_in_stage": ln.days_in_stage or 0,
            "sla_status": ln.sla_status or "unknown",
            "risk_score": ln.risk_score or 0,
            "lock_expired": lock_expired,
            "lock_expiration_date": ln.lock_expiration_date,
            "closing_date": ln.closing_date,
            "funded_date": ln.funded_date,
            "loan_officer": ln.loan_officer_name or "",
            "created_at": ln.created_at,
        })

    return rows


COMPLIANCE_COLUMNS = [
    "loan_id", "loan_number", "borrower", "stage", "loan_type", "amount",
    "days_in_stage", "sla_status", "risk_score", "lock_expired",
    "lock_expiration_date", "closing_date", "funded_date",
    "loan_officer", "created_at",
]


# =============================================================================
# Module-Level Convenience
# =============================================================================

class ReportExporter:
    """Convenience class aggregating all export functions."""
    generate_pdf = staticmethod(generate_pdf)
    generate_excel = staticmethod(generate_excel)
    generate_csv = staticmethod(generate_csv)
    generate_sla_compliance_report = staticmethod(generate_sla_compliance_report)
    export_to_csv = staticmethod(export_to_csv)
    export_to_excel = staticmethod(export_to_excel)
    format_pipeline_report = staticmethod(format_pipeline_report)
    format_production_report = staticmethod(format_production_report)
    format_compliance_report = staticmethod(format_compliance_report)


report_exporter = ReportExporter()
