"""
Bank Statement Worksheet Generator

Generates Excel worksheets in the Non-QM Connect format with extracted bank statement data.
Creates multi-tab workbook with Personal, Business Method 1/2/3 sheets.
"""

import io
import os
import logging
from datetime import datetime, date
from decimal import Decimal
from typing import Optional, Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class BankStatementWorksheetGenerator:
    """Generates Excel worksheets for Non-QM bank statement income calculation."""

    # Style definitions
    HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
    INPUT_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Yellow
    TOTAL_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Green
    QUALIFYING_FILL = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    QUALIFYING_FONT = Font(color="FFFFFF", bold=True, size=14)

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    MONTHS = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    def __init__(self):
        pass

    def generate_worksheet(
        self,
        borrower_name: str,
        business_name: Optional[str],
        personal_data: Optional[List[Dict]] = None,
        business_data: Optional[List[Dict]] = None,
        calculation_method: str = "UNIFORM_EXPENSE_RATIO",
        ownership_percentage: float = 100.0,
        ltv_percentage: float = 80.0,
        cpa_expense_ratio: Optional[float] = None,
        months: int = 12
    ) -> io.BytesIO:
        """
        Generate a complete bank statement worksheet.

        Args:
            borrower_name: Borrower's full name
            business_name: Business name (if applicable)
            personal_data: List of monthly personal statement data
            business_data: List of monthly business statement data
            calculation_method: Income calculation method
            ownership_percentage: Business ownership percentage
            ltv_percentage: Loan-to-value percentage
            cpa_expense_ratio: CPA-stated expense ratio (for Method 3)
            months: Number of months (12 or 24)

        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Input Tab
        self._create_input_tab(wb, borrower_name, business_name, months)

        # Create Personal Statements tab
        self._create_personal_tab(
            wb, borrower_name, personal_data or [],
            has_separate_business=business_data is not None and len(business_data) > 0,
            months=months
        )

        # Create Business tabs (all three methods)
        self._create_business_method1_tab(
            wb, borrower_name, business_name, business_data or [],
            ownership_percentage, ltv_percentage, months
        )

        self._create_business_method2_tab(
            wb, borrower_name, business_name, business_data or [],
            ownership_percentage, months
        )

        self._create_business_method3_tab(
            wb, borrower_name, business_name, business_data or [],
            ownership_percentage, cpa_expense_ratio, months
        )

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _create_input_tab(
        self,
        wb: Workbook,
        borrower_name: str,
        business_name: Optional[str],
        months: int
    ):
        """Create the Input Tab sheet."""
        ws = wb.create_sheet("Input Tab")

        # Title
        ws.merge_cells('A6:F6')
        ws['A6'] = "Bank Statement Worksheet (Three Methods)"
        ws['A6'].font = Font(bold=True, size=16)

        # Instructions
        ws['B7'] = "Fill in Yellow Cells"
        ws['B7'].fill = self.INPUT_FILL

        # Borrower Info
        ws['B8'] = "Borrower Name:"
        ws['C8'] = borrower_name
        ws['C8'].fill = self.INPUT_FILL

        ws['B9'] = "Bank Name:"
        ws['C9'].fill = self.INPUT_FILL

        ws['B10'] = "Last Four of Account:"
        ws['C10'].fill = self.INPUT_FILL

        # Business Info
        ws['C11'] = "Business Name"
        ws['D11'] = business_name or ""
        ws['D11'].fill = self.INPUT_FILL

        ws['C12'] = "Months of Statements"
        ws['D12'] = months
        ws['D12'].fill = self.INPUT_FILL

        ws['C13'] = "Percentage of Business Owned"
        ws['D13'].fill = self.INPUT_FILL

        ws['D14'] = "MOST RECENT MONTH STATEMENT IN FILE"
        ws['C15'] = "Date (MM/DD/YYYY):"
        ws['D15'].fill = self.INPUT_FILL

        # Statement headers
        ws['B16'] = "MOST RECENT 12 MONTHS OF STATEMENTS-ORDER NEWEST TO OLDEST"
        ws['B16'].font = Font(bold=True)

        headers = ["", "Statement Ending", "End Date", "Total Deposits",
                   "Ineligible Deposits", "Eligible Deposits", "Beginning Balance",
                   "Ending Balance", "NSFs"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=17, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL if header else None
            cell.font = self.HEADER_FONT if header else None

        # Month rows
        for i in range(months):
            row = 18 + i
            ws.cell(row=row, column=1).value = i
            ws.cell(row=row, column=2).value = self.MONTHS[(12 - i - 1) % 12] if i < 12 else self.MONTHS[(24 - i - 1) % 12]

            # Input cells (yellow)
            for col in range(3, 10):
                cell = ws.cell(row=row, column=col)
                cell.fill = self.INPUT_FILL
                if col in [4, 5, 6, 7, 8]:  # Money columns
                    cell.number_format = '$#,##0.00'
                cell.border = self.THIN_BORDER

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 8

    def _create_personal_tab(
        self,
        wb: Workbook,
        borrower_name: str,
        data: List[Dict],
        has_separate_business: bool,
        months: int
    ):
        """Create the Personal Statements sheet."""
        ws = wb.create_sheet("Personal Statements")

        # Title
        ws.merge_cells('B6:H6')
        ws['B6'] = "Personal Bank Statements (Separate from Business Account)"
        ws['B6'].font = Font(bold=True, size=14)

        ws['B7'] = "Fill in Input Tab First-Fill in Yellow Cells Second"
        ws['B7'].fill = self.INPUT_FILL

        # Info rows
        ws['B8'] = "Borrower Name:"
        ws['D8'] = borrower_name

        ws['B9'] = "Bank Name:"
        ws['D9'] = data[0].get("bank_name", "") if data else ""

        ws['C10'] = "Months of Statements"
        ws['D10'] = months

        ws['B11'] = "Last Four of Account:"
        ws['D11'] = data[0].get("account_last_four", "") if data else ""

        ws['A12'] = "Does Borrower Have a Separate Biz Account:"
        ws['D12'] = "YES" if has_separate_business else "NO"

        # Headers
        ws['B13'] = "Most Recent 12 Months"
        ws['B13'].font = Font(bold=True)

        headers = ["", "Statement Ending", "End Date", "Total Deposits",
                   "Ineligible Deposits", "Eligible Deposits", "Beginning Balance",
                   "Ending Balance", "NSFs"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=14, column=col)
            cell.value = header
            cell.font = Font(bold=True) if header else None
            if header:
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

        # Data rows
        for i in range(months):
            row = 15 + i
            month_data = data[i] if i < len(data) else {}

            ws.cell(row=row, column=1).value = ""
            ws.cell(row=row, column=2).value = month_data.get("statement_month", self.MONTHS[(12 - i - 1) % 12])
            ws.cell(row=row, column=3).value = month_data.get("statement_end_date", "")
            ws.cell(row=row, column=4).value = month_data.get("total_deposits", 0)
            ws.cell(row=row, column=5).value = month_data.get("ineligible_deposits_total", 0)
            ws.cell(row=row, column=6).value = month_data.get("eligible_deposits_total", 0)
            ws.cell(row=row, column=7).value = month_data.get("beginning_balance", 0)
            ws.cell(row=row, column=8).value = month_data.get("ending_balance", 0)
            ws.cell(row=row, column=9).value = month_data.get("nsf_count", 0)

            # Format
            for col in range(4, 9):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '$#,##0.00'
                cell.border = self.THIN_BORDER

        # Totals row
        total_row = 15 + months
        ws.cell(row=total_row, column=2).value = "TOTALS"
        ws.cell(row=total_row, column=2).font = Font(bold=True)

        for col in [4, 5, 6]:
            cell = ws.cell(row=total_row, column=col)
            cell.value = f"=SUM({get_column_letter(col)}15:{get_column_letter(col)}{total_row-1})"
            cell.number_format = '$#,##0.00'
            cell.fill = self.TOTAL_FILL
            cell.font = Font(bold=True)

        # Set column widths
        for col, width in enumerate([5, 18, 12, 15, 18, 18, 18, 15, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_business_method1_tab(
        self,
        wb: Workbook,
        borrower_name: str,
        business_name: Optional[str],
        data: List[Dict],
        ownership_pct: float,
        ltv_pct: float,
        months: int
    ):
        """Create Business-Method #1 (Uniform Expense Ratio) sheet."""
        ws = wb.create_sheet("Business-Method #1")

        # Title
        ws.merge_cells('B6:H6')
        ws['B6'] = "Business Bank Statements - Uniform Expense Ratio (Method One)"
        ws['B6'].font = Font(bold=True, size=14)

        ws['B7'] = "Fill in Input Tab First-Fill in Yellow Cells Second"

        # Info
        ws['B8'] = "Borrower Name:"
        ws['D8'] = borrower_name

        ws['C9'] = "Business Name:"
        ws['D9'] = business_name or ""

        ws['B10'] = "Bank Name:"
        ws['D10'] = data[0].get("bank_name", "") if data else ""

        ws['C11'] = "Months of Statements:"
        ws['D11'] = months

        ws['B12'] = "Last Four of Account:"
        ws['D12'] = data[0].get("account_last_four", "") if data else ""

        ws['B13'] = "Borrower's Ownership (%):"
        ws['E13'] = ownership_pct
        ws['E13'].number_format = '0.00%'

        # Expense ratio note
        expense_ratio = 0.50 if ltv_pct <= 80 else 0.60
        ws['K10'] = "<=80%"
        ws['K11'] = ">80%"

        # Headers
        ws['B14'] = "Most Recent 12 Months"
        headers = ["", "Statement Ending", "End Date", "Total Deposits",
                   "Ineligible Deposits", "Eligible Deposits", "Beginning Balance",
                   "Ending Balance", "NSFs"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=15, column=col)
            cell.value = header
            if header:
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

        # Data rows
        for i in range(months):
            row = 16 + i
            month_data = data[i] if i < len(data) else {}

            ws.cell(row=row, column=2).value = month_data.get("statement_month", self.MONTHS[(12 - i - 1) % 12])
            ws.cell(row=row, column=3).value = month_data.get("statement_end_date", "")
            ws.cell(row=row, column=4).value = month_data.get("total_deposits", 0)
            ws.cell(row=row, column=5).value = month_data.get("ineligible_deposits_total", 0)
            ws.cell(row=row, column=6).value = month_data.get("eligible_deposits_total", 0)
            ws.cell(row=row, column=7).value = month_data.get("beginning_balance", 0)
            ws.cell(row=row, column=8).value = month_data.get("ending_balance", 0)
            ws.cell(row=row, column=9).value = month_data.get("nsf_count", 0)

            for col in range(4, 9):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '$#,##0.00'
                cell.border = self.THIN_BORDER

        # Totals and calculations
        total_row = 16 + months
        ws.cell(row=total_row, column=2).value = "TOTALS"
        ws.cell(row=total_row, column=2).font = Font(bold=True)

        for col in [4, 5, 6]:
            cell = ws.cell(row=total_row, column=col)
            cell.value = f"=SUM({get_column_letter(col)}16:{get_column_letter(col)}{total_row-1})"
            cell.number_format = '$#,##0.00'
            cell.fill = self.TOTAL_FILL

        # Income calculation section
        calc_row = total_row + 3
        ws.cell(row=calc_row, column=2).value = "INCOME CALCULATION"
        ws.cell(row=calc_row, column=2).font = Font(bold=True, size=12)

        ws.cell(row=calc_row+1, column=2).value = "Total Eligible Deposits:"
        ws.cell(row=calc_row+1, column=4).value = f"=F{total_row}"
        ws.cell(row=calc_row+1, column=4).number_format = '$#,##0.00'

        ws.cell(row=calc_row+2, column=2).value = "Expense Ratio:"
        ws.cell(row=calc_row+2, column=4).value = expense_ratio
        ws.cell(row=calc_row+2, column=4).number_format = '0%'

        ws.cell(row=calc_row+3, column=2).value = "Net Income:"
        ws.cell(row=calc_row+3, column=4).value = f"=D{calc_row+1}*(1-D{calc_row+2})"
        ws.cell(row=calc_row+3, column=4).number_format = '$#,##0.00'

        ws.cell(row=calc_row+4, column=2).value = "Ownership %:"
        ws.cell(row=calc_row+4, column=4).value = ownership_pct / 100
        ws.cell(row=calc_row+4, column=4).number_format = '0%'

        ws.cell(row=calc_row+5, column=2).value = "Owner's Share:"
        ws.cell(row=calc_row+5, column=4).value = f"=D{calc_row+3}*D{calc_row+4}"
        ws.cell(row=calc_row+5, column=4).number_format = '$#,##0.00'

        ws.cell(row=calc_row+7, column=2).value = "MONTHLY QUALIFYING INCOME:"
        ws.cell(row=calc_row+7, column=2).font = Font(bold=True)
        ws.cell(row=calc_row+7, column=4).value = f"=D{calc_row+5}/{months}"
        ws.cell(row=calc_row+7, column=4).number_format = '$#,##0.00'
        ws.cell(row=calc_row+7, column=4).fill = self.QUALIFYING_FILL
        ws.cell(row=calc_row+7, column=4).font = self.QUALIFYING_FONT

        # Column widths
        for col, width in enumerate([5, 18, 12, 15, 18, 18, 18, 15, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_business_method2_tab(
        self,
        wb: Workbook,
        borrower_name: str,
        business_name: Optional[str],
        data: List[Dict],
        ownership_pct: float,
        months: int
    ):
        """Create Business-Method #2 (CPA P&L) sheet."""
        ws = wb.create_sheet("Business-Method #2")

        ws.merge_cells('B6:H6')
        ws['B6'] = "Business Bank Statements - Utilize CPA Prepared P&L (Method Two)"
        ws['B6'].font = Font(bold=True, size=14)

        ws['B7'] = "Fill in Input Tab First-Fill in Yellow Cells Second"

        ws['B8'] = "Borrower Name:"
        ws['D8'] = borrower_name

        ws['C9'] = "Business Name:"
        ws['D9'] = business_name or ""

        ws['B10'] = "Bank Name:"
        ws['D10'] = data[0].get("bank_name", "") if data else ""

        ws['C11'] = "Number of Months of Statements:"
        ws['D11'] = months

        ws['B12'] = "Last Four of Account:"
        ws['D12'] = data[0].get("account_last_four", "") if data else ""

        ws['B13'] = "Borrower's Ownership (%):"
        ws['E13'] = ownership_pct
        ws['E13'].number_format = '0.00%'

        # Same structure as Method 1 for the data table
        headers = ["", "Statement Ending", "End Date", "Total Deposits",
                   "Ineligible Deposits", "Eligible Deposits", "Beginning Balance",
                   "Ending Balance", "NSFs"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=15, column=col)
            cell.value = header
            if header:
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

        for i in range(months):
            row = 16 + i
            month_data = data[i] if i < len(data) else {}

            ws.cell(row=row, column=2).value = month_data.get("statement_month", self.MONTHS[(12 - i - 1) % 12])
            ws.cell(row=row, column=3).value = month_data.get("statement_end_date", "")
            ws.cell(row=row, column=4).value = month_data.get("total_deposits", 0)
            ws.cell(row=row, column=5).value = month_data.get("ineligible_deposits_total", 0)
            ws.cell(row=row, column=6).value = month_data.get("eligible_deposits_total", 0)
            ws.cell(row=row, column=7).value = month_data.get("beginning_balance", 0)
            ws.cell(row=row, column=8).value = month_data.get("ending_balance", 0)
            ws.cell(row=row, column=9).value = month_data.get("nsf_count", 0)

            for col in range(4, 9):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '$#,##0.00'
                cell.border = self.THIN_BORDER

        # CPA P&L Section
        pl_row = 16 + months + 3
        ws.cell(row=pl_row, column=2).value = "CPA PREPARED P&L DATA"
        ws.cell(row=pl_row, column=2).font = Font(bold=True)

        ws.cell(row=pl_row+1, column=2).value = "Gross Revenue:"
        ws.cell(row=pl_row+1, column=4).fill = self.INPUT_FILL
        ws.cell(row=pl_row+1, column=4).number_format = '$#,##0.00'

        ws.cell(row=pl_row+2, column=2).value = "Total Expenses:"
        ws.cell(row=pl_row+2, column=4).fill = self.INPUT_FILL
        ws.cell(row=pl_row+2, column=4).number_format = '$#,##0.00'

        ws.cell(row=pl_row+3, column=2).value = "Net Income (from P&L):"
        ws.cell(row=pl_row+3, column=4).value = f"=D{pl_row+1}-D{pl_row+2}"
        ws.cell(row=pl_row+3, column=4).number_format = '$#,##0.00'

        for col, width in enumerate([5, 18, 12, 15, 18, 18, 18, 15, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_business_method3_tab(
        self,
        wb: Workbook,
        borrower_name: str,
        business_name: Optional[str],
        data: List[Dict],
        ownership_pct: float,
        cpa_expense_ratio: Optional[float],
        months: int
    ):
        """Create Business-Method #3 (CPA Letter Expense Ratio) sheet."""
        ws = wb.create_sheet("Business-Method #3")

        ws.merge_cells('B6:H6')
        ws['B6'] = "Business Bank Statements - Utilize Expense Ratio in CPA Letter (Method Three)"
        ws['B6'].font = Font(bold=True, size=14)

        ws['B7'] = "Fill in Input Tab First-Fill in Yellow Cells Second"

        ws['B8'] = "Borrower Name:"
        ws['D8'] = borrower_name

        ws['C9'] = "Business Name:"
        ws['D9'] = business_name or ""

        ws['B10'] = "Bank Name:"
        ws['D10'] = data[0].get("bank_name", "") if data else ""

        ws['C11'] = "Months of Statements:"
        ws['D11'] = months

        ws['B12'] = "Last Four of Account:"
        ws['D12'] = data[0].get("account_last_four", "") if data else ""

        ws['B13'] = "CPA Stated Expense Ratio:"
        ws['D13'] = (cpa_expense_ratio or 50) / 100
        ws['D13'].number_format = '0%'
        ws['D13'].fill = self.INPUT_FILL

        ws['B14'] = "Borrower's Ownership (%):"
        ws['E14'] = ownership_pct
        ws['E14'].number_format = '0.00%'

        # Data table
        headers = ["", "Statement Ending", "End Date", "Total Deposits",
                   "Ineligible Deposits", "Eligible Deposits", "Beginning Balance",
                   "Ending Balance", "NSFs"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=16, column=col)
            cell.value = header
            if header:
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

        for i in range(months):
            row = 17 + i
            month_data = data[i] if i < len(data) else {}

            ws.cell(row=row, column=2).value = month_data.get("statement_month", self.MONTHS[(12 - i - 1) % 12])
            ws.cell(row=row, column=3).value = month_data.get("statement_end_date", "")
            ws.cell(row=row, column=4).value = month_data.get("total_deposits", 0)
            ws.cell(row=row, column=5).value = month_data.get("ineligible_deposits_total", 0)
            ws.cell(row=row, column=6).value = month_data.get("eligible_deposits_total", 0)
            ws.cell(row=row, column=7).value = month_data.get("beginning_balance", 0)
            ws.cell(row=row, column=8).value = month_data.get("ending_balance", 0)
            ws.cell(row=row, column=9).value = month_data.get("nsf_count", 0)

            for col in range(4, 9):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '$#,##0.00'
                cell.border = self.THIN_BORDER

        for col, width in enumerate([5, 18, 12, 15, 18, 18, 18, 15, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = width


# Singleton
_generator = None


def get_worksheet_generator() -> BankStatementWorksheetGenerator:
    """Get or create the worksheet generator singleton."""
    global _generator
    if _generator is None:
        _generator = BankStatementWorksheetGenerator()
    return _generator
